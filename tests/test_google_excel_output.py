import sys, os, tempfile, json
from datetime import datetime
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pandas as pd
from openpyxl import load_workbook
from src.excel_exporter import export_google_report


def test_export_google_report_creates_file():
    df = pd.DataFrame({
        "原始關鍵字": ["貼紙印刷", "標籤印刷"],
        "平均每月搜尋量": [1000, 500],
        "競爭程度": ["低", "中"],
    })
    query_info = {
        "query_time": "2026-07-25 12:00:00",
        "geo": "台灣",
        "language": "中文 (繁體)",
        "network": "Google Search",
        "customer_id": "1234567890",
        "from_cache": False,
        "seed_keywords": ["貼紙印刷", "標籤印刷"],
        "avg_volume": "750",
    }
    with tempfile.TemporaryDirectory() as tmpdir:
        filepath = export_google_report(df, tmpdir, query_info)
        assert os.path.exists(filepath), "Excel 檔案應存在"
        wb = load_workbook(filepath)
        sheet_names = wb.sheetnames
        assert "查詢摘要" in sheet_names, "應有查詢摘要工作表"
        assert "Google 關鍵字結果" in sheet_names, "應有 Google 關鍵字結果工作表"
        print("  PASS: Excel 檔案建立且工作表正確")


def test_export_google_report_summary_content():
    df = pd.DataFrame({
        "原始關鍵字": ["貼紙印刷"],
        "平均每月搜尋量": [1000],
        "競爭程度": ["低"],
    })
    query_info = {
        "query_time": "2026-07-25 12:00:00",
        "geo": "台灣",
        "language": "中文 (繁體)",
        "network": "Google Search",
        "customer_id": "1234567890",
        "from_cache": True,
        "seed_keywords": ["貼紙印刷"],
        "avg_volume": "1000",
    }
    with tempfile.TemporaryDirectory() as tmpdir:
        filepath = export_google_report(df, tmpdir, query_info)
        wb = load_workbook(filepath)
        ws = wb["查詢摘要"]
        cell_map = {}
        for row in ws.iter_rows(min_row=3, values_only=False):
            if row[0].value:
                cell_map[row[0].value] = row[1].value
        assert cell_map.get("使用快取") == "是"
        assert cell_map.get("Customer ID") == "1234567890"
        assert cell_map.get("地區") == "台灣"
        assert cell_map.get("語言") == "中文 (繁體)"
        assert cell_map.get("搜尋網路") == "Google Search"
        print("  PASS: 查詢摘要內容正確")


def test_export_google_report_data_rows():
    df = pd.DataFrame({
        "原始關鍵字": ["貼紙印刷", "標籤印刷"],
        "平均每月搜尋量": [1000, 500],
        "競爭程度": ["低", "中"],
    })
    query_info = {
        "query_time": "2026-07-25 12:00:00",
        "geo": "台灣",
        "language": "中文 (繁體)",
        "network": "Google Search",
        "customer_id": "1234567890",
        "from_cache": False,
        "seed_keywords": ["貼紙印刷"],
        "avg_volume": "750",
    }
    with tempfile.TemporaryDirectory() as tmpdir:
        filepath = export_google_report(df, tmpdir, query_info)
        wb = load_workbook(filepath)
        ws = wb["Google 關鍵字結果"]
        assert ws.cell(1, 1).value == "原始關鍵字"
        assert ws.cell(2, 1).value == "貼紙印刷"
        assert ws.cell(3, 1).value == "標籤印刷"
        assert ws.cell(2, 2).value == 1000
        print("  PASS: Google 關鍵字結果資料正確")


def test_export_google_report_filename():
    df = pd.DataFrame({"原始關鍵字": ["測試"]})
    query_info = {"query_time": "2026-07-25 12:00:00", "seed_keywords": ["測試"], "from_cache": False}
    with tempfile.TemporaryDirectory() as tmpdir:
        filepath = export_google_report(df, tmpdir, query_info)
        assert "Google關鍵字查詢_" in os.path.basename(str(filepath))
        assert str(filepath).endswith(".xlsx")
        print("  PASS: 檔案名稱格式正確")


def run_all_tests():
    passed = failed = 0
    tests = [
        ("Excel 建立與工作表", test_export_google_report_creates_file),
        ("查詢摘要內容", test_export_google_report_summary_content),
        ("關鍵字結果資料", test_export_google_report_data_rows),
        ("檔案名稱", test_export_google_report_filename),
    ]
    for name, func in tests:
        try:
            func()
            passed += 1
        except Exception as e:
            print(f"  FAIL: {e}")
            import traceback; traceback.print_exc()
            failed += 1
        print()
    print(f"測試結果: {passed} 通過, {failed} 失敗, 共 {len(tests)} 項")
    return failed == 0


if __name__ == "__main__":
    success = run_all_tests()
    sys.exit(0 if success else 1)
