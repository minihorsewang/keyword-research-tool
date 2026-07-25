import logging
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HIGH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LOW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
IRRELEVANT_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def auto_width(ws, min_width=8, max_width=40):
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def apply_header(ws, headers):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_priority_color(ws, row_idx, priority, max_col):
    if priority == "高":
        fill = HIGH_FILL
    elif priority == "中":
        fill = MEDIUM_FILL
    elif priority == "低":
        fill = LOW_FILL
    else:
        fill = IRRELEVANT_FILL
    for col_idx in range(1, max_col + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def write_sheet_summary(ws, summary_data):
    ws.cell(row=1, column=1, value="分析摘要").font = Font(bold=True, size=14)
    for idx, (key, val) in enumerate(summary_data.items(), 3):
        ws.cell(row=idx, column=1, value=key).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=val)
    auto_width(ws)


def write_sheet_detail(ws, df, columns, sheet_name=""):
    apply_header(ws, columns)
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col in enumerate(columns, 1):
            val = row.get(col, "")
            if isinstance(val, float):
                val = round(val, 2)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
        priority = row.get("優先級", "")
        apply_priority_color(ws, row_idx, priority, len(columns))
    auto_width(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_excel(df, clusters, pages, output_dir, summary_data):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"印刷關鍵字分析_{timestamp}.xlsx"
    filepath = Path(output_dir) / filename

    wb = Workbook()

    # 1. 分析摘要
    ws_summary = wb.active
    ws_summary.title = "分析摘要"
    write_sheet_summary(ws_summary, summary_data)

    # 2. 關鍵字總表
    ws_keywords = wb.create_sheet("關鍵字總表")
    detail_columns = [
        "原始關鍵字", "標準化關鍵字", "平均每月搜尋量", "競爭程度", "競爭指數",
        "頁首出價低", "頁首出價高", "三個月變化", "年增率",
        "產品大類", "材質", "用途", "加工", "搜尋意圖", "地區",
        "是否可能無關", "是否完全重複", "是否高度相似",
        "搜尋量分數", "交易意圖分數", "承接能力", "毛利潛力", "競爭難度",
        "機會分數", "標準化分數", "優先級", "評分原因", "群組名稱", "建議頁面類型"
    ]
    # 將分群與頁面建議合併回 df
    df_out = df.copy()
    df_out["群組名稱"] = ""
    df_out["建議頁面類型"] = ""
    for c in clusters:
        for kw in c["次要關鍵字"].split("、"):
            if kw:
                mask = df_out["標準化關鍵字"] == kw
                df_out.loc[mask, "群組名稱"] = c["群組名稱"]
                df_out.loc[mask, "建議頁面類型"] = c["建議頁面"]
        primary = c["主關鍵字"]
        mask = df_out["標準化關鍵字"] == primary
        df_out.loc[mask, "群組名稱"] = c["群組名稱"]
        df_out.loc[mask, "建議頁面類型"] = c["建議頁面"]
    write_sheet_detail(ws_keywords, df_out, detail_columns)

    # 3. 高優先關鍵字
    ws_high = wb.create_sheet("高優先關鍵字")
    high_df = df[(df["優先級"] == "高") & (~df["是否可能無關"])]
    high_columns = [c for c in detail_columns if c not in ["是否完全重複", "是否高度相似"]]
    write_sheet_detail(ws_high, high_df, high_columns)

    # 4. 關鍵字分群
    ws_clusters = wb.create_sheet("關鍵字分群")
    cluster_columns = ["群組名稱", "主關鍵字", "次要關鍵字", "群組搜尋量", "群組意圖", "建議頁面", "優先級"]
    apply_header(ws_clusters, cluster_columns)
    for row_idx, c in enumerate(clusters, 2):
        for col_idx, col in enumerate(cluster_columns, 1):
            val = c.get(col, "")
            cell = ws_clusters.cell(row=row_idx, column=col_idx, value=val)
        priority = c.get("優先級", "")
        apply_priority_color(ws_clusters, row_idx, priority, len(cluster_columns))
    auto_width(ws_clusters)
    ws_clusters.freeze_panes = "A2"
    ws_clusters.auto_filter.ref = ws_clusters.dimensions

    # 5. 網站頁面規劃
    ws_pages = wb.create_sheet("網站頁面規劃")
    page_columns = ["頁面名稱", "頁面類型", "主關鍵字", "次要關鍵字", "Slug", "H1", "內容方向", "CTA", "優先級", "狀態", "備註"]
    apply_header(ws_pages, page_columns)
    for row_idx, p in enumerate(pages, 2):
        for col_idx, col in enumerate(page_columns, 1):
            val = p.get(col, "")
            cell = ws_pages.cell(row=row_idx, column=col_idx, value=val)
        priority = p.get("優先級", "")
        apply_priority_color(ws_pages, row_idx, priority, len(page_columns))
    auto_width(ws_pages)
    ws_pages.freeze_panes = "A2"
    ws_pages.auto_filter.ref = ws_pages.dimensions

    # 6. 可能無關
    ws_irr = wb.create_sheet("可能無關")
    irr_df = df[df["是否可能無關"]]
    irr_columns = ["原始關鍵字", "標準化關鍵字", "平均每月搜尋量", "搜尋意圖"]
    write_sheet_detail(ws_irr, irr_df, irr_columns)

    # 7. 原始資料
    ws_raw = wb.create_sheet("原始資料")
    raw_columns = list(df.columns)
    # remove semicolons properly
    write_sheet_detail(ws_raw, df, raw_columns)

    # 8. 執行紀錄
    ws_log = wb.create_sheet("執行紀錄")
    log_entries = [
        ("執行時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("輸入檔案", summary_data.get("輸入檔案", "")),
        ("程式版本", "1.0.0"),
        ("處理數量", summary_data.get("有效關鍵字數", 0)),
        ("錯誤與警告", ""),
        ("使用的設定檔版本", "1.0.0"),
    ]
    for idx, (key, val) in enumerate(log_entries, 1):
        ws_log.cell(row=idx, column=1, value=key).font = Font(bold=True)
        ws_log.cell(row=idx, column=2, value=val)
    auto_width(ws_log)

    # 設定頁面
    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = "4472C4"

    wb.save(filepath)
    logger.info(f"Excel 已匯出: {filepath}")
    return filepath
